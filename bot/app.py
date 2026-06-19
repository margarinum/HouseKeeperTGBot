from __future__ import annotations

import html
import logging
from contextlib import asynccontextmanager
from logging.handlers import TimedRotatingFileHandler
import os
import re
from typing import Dict, List
from io import BytesIO

from openpyxl import Workbook

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ChatType, ParseMode
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, Message, ReplyKeyboardRemove, BufferedInputFile, FSInputFile
from telethon import TelegramClient

from .config import Settings, load_settings
from .barrier_client import BARRIERS, BarrierClient, BarrierClientError, ChangeDeviceRequired
from .gas_client import GasClient, GasClientError
from .deepseek_client import DeepSeekClient
from .keyboards import (
    poll_vote_keyboard,
    poll_manage_keyboard,
    docs_keyboard,
    user_docs_keyboard,
    faq_admin_keyboard,
    faq_user_keyboard,
    ai_conversation_keyboard,
    DOCS_BTN,
    FAQ_BTN,
    AI_BTN,
    ADMIN_BTN,
    CANCEL_BTN,
    DELETE_VERIFY_BTN,
    VERIFY_BTN,
    admin_menu,
    main_menu,
    houses_menu,
    unverified_actions_menu,
    verified_actions_menu,
    paid_actions_menu,
    barriers_menu,
    confirm_revoke_menu,
)
from .membership import MembershipManager
from .prompt_log import log_user_ai_exchange, setup_prompt_logger
from .storage import Storage

logger = logging.getLogger(__name__)
router = Router()


class VerifyStates(StatesGroup):
    waiting_for_house = State()
    waiting_for_entrance = State()
    waiting_for_floor = State()
    waiting_for_apartment = State()


class AdminStates(StatesGroup):
    waiting_for_broadcast_text = State()
    waiting_for_add_admin = State()
    waiting_for_remove_admin = State()
    waiting_for_revoke_user_id = State()
    waiting_for_reset_attempts_user_id = State()
    waiting_for_remove_limit_house = State()
    waiting_for_remove_limit_apartment = State()
    waiting_for_broadcast_message = State()
    waiting_for_doc_name = State()
    waiting_for_doc_file = State()
    waiting_for_faq_text = State()
    waiting_for_ai_question = State()
    waiting_for_ai_choice = State()
    waiting_for_poll_question = State()
    waiting_for_poll_options = State()
    waiting_for_manual_verify_user_id = State()
    waiting_for_manual_verify_house = State()
    waiting_for_manual_verify_apartment = State()
    waiting_for_barrier_pin = State()


class AppContext:
    def __init__(self, settings: Settings):
        self.ai: DeepSeekClient | None = None
        self.settings = settings
        self.storage = Storage(settings.db_path)
        self.gas = GasClient(settings.gas_web_app_url, settings.gas_shared_secret)
        self.barrier = BarrierClient(
            settings.barrier_api_url,
            settings.barrier_login,
            settings.barrier_password,
            settings.barrier_config_path,
        )
        self.bot = Bot(settings.bot_token, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
        self.bot_user_id: int | None = None
        self.dp = Dispatcher()
        self.dp.include_router(router)
        self.telethon = TelegramClient(settings.telethon_session, settings.api_id, settings.api_hash)
        self.membership = MembershipManager(self.telethon, settings.group_id)


app: AppContext | None = None


def is_private_message(message: Message) -> bool:
    return bool(message.chat and message.chat.type == ChatType.PRIVATE)


def is_private_callback(callback: CallbackQuery) -> bool:
    return bool(callback.message and callback.message.chat.type == ChatType.PRIVATE)


def message_user_display_name(message: Message) -> str:
    full_name = message.from_user.full_name.strip() or str(message.from_user.id)
    if message.from_user.username:
        return f"{full_name} (@{message.from_user.username})"
    return full_name


def mention_html(user_id: int, username: str | None, full_name: str) -> str:
    if username:
        return f"@{html.escape(username)}"
    label = html.escape(full_name or str(user_id))
    return f"<a href='tg://user?id={user_id}'>{label}</a>"


async def get_live_display_name(user_id: int) -> tuple[str | None, str]:
    """Fetch real first+last name and username from Telegram (not from local DB)."""
    assert app is not None
    try:
        member = await app.bot.get_chat_member(app.settings.group_id, user_id)
        user = member.user
        full_name = (user.full_name or "").strip() or str(user_id)
        return user.username, full_name
    except Exception:
        logger.debug("Could not fetch live name for user_id=%s", user_id)
        return None, str(user_id)


async def mention_html_live(user_id: int) -> str:
    """Build mention using real Telegram name."""
    username, full_name = await get_live_display_name(user_id)
    return mention_html(user_id, username, full_name)


def valid_house_input(value: str) -> bool:
    text = (value or "").strip()
    return bool(text and text == text.lower() and re.fullmatch(r"[а-яё0-9\s.\-]+", text))


def user_mention_from_aiogram(user) -> str:
    username = getattr(user, "username", None)
    user_id = getattr(user, "id", None)
    full_name = getattr(user, "full_name", "") or str(user_id)

    if username:
        return f"@{html.escape(username)}"

    return f"<a href='tg://user?id={user_id}'>{html.escape(full_name)}</a>"


def new_member_instruction_text() -> str:
    return (
        "Добро пожаловать в чат сообщества домов.\n\n"
        "Для доступа к общению нужно пройти верификацию, в противном случае вы можете быть ограничены в отправке сообщений.\n"
        "Напишите боту (мне) в личные сообщения и нажмите «Верифицироваться».\n"
    )


async def send_base_error(message: Message) -> None:
    await message.answer("Ошибка связи с базой. Попробуйте позже или обратитесь к администратору.")


async def send_verification_error(message: Message, text: str | None = None) -> None:
    await message.answer(text or "Не удалось завершить верификацию. Попробуйте позже или обратитесь к администратору.")


async def reject_group_interaction(message: Message, text: str | None = None) -> None:
    await message.answer(text or "Для верификации напишите мне в личные сообщения.", reply_markup=ReplyKeyboardRemove())


async def reject_group_callback(callback: CallbackQuery) -> None:
    await callback.answer("Эта функция доступна только в личных сообщениях с ботом.", show_alert=True)


async def is_admin(user_id: int) -> bool:
    assert app is not None
    return await app.storage.is_admin(user_id)


async def build_main_menu(user_id: int):
    assert app is not None
    verified = await app.storage.get_verification(user_id)
    return main_menu(bool(verified), await is_admin(user_id))


async def sync_group_members() -> None:
    assert app is not None
    members = await _group_human_members()
    present_ids: List[int] = []
    for member in members:
        present_ids.append(member.user_id)
        await app.storage.touch_user(member.user_id, member.username, member.full_name)
    await app.storage.mark_users_absent_except(present_ids)
    logger.info("Membership sync complete. Human members in group: %s", len(present_ids))


async def _ensure_bot_user_id() -> int | None:
    assert app is not None
    if app.bot_user_id is None:
        me = await app.bot.get_me()
        app.bot_user_id = me.id
    return app.bot_user_id


async def _group_human_members() -> list:
    """Участники группы: люди, без Telegram-бота сервиса."""
    assert app is not None
    bot_id = await _ensure_bot_user_id()
    members = await app.membership.list_all_human_members()
    if bot_id is None:
        return members
    return [m for m in members if m.user_id != bot_id]


def _gas_user_id(item: dict) -> int | None:
    raw = str(item.get("user_id", "")).strip()
    return int(raw) if raw.isdigit() else None


def _gas_row_is_paid(item: dict) -> bool:
    return str(item.get("payed", "0")).strip() == "1"


async def _table_payment_user_ids() -> tuple[set[int], set[int]]:
    """(все user_id из таблицы, user_id с payed=1 хотя бы в одной строке)."""
    assert app is not None
    app.gas.invalidate_cache()
    table_users = await app.gas.list_users()
    all_ids: set[int] = set()
    paid_ids: set[int] = set()
    for item in table_users:
        user_id = _gas_user_id(item)
        if user_id is None:
            continue
        all_ids.add(user_id)
        if _gas_row_is_paid(item):
            paid_ids.add(user_id)
    return all_ids, paid_ids


async def _live_membership_verification() -> tuple[set[int], set[int], list]:
    """Участники группы (без бота) и верифицированные среди них; админы учитываются."""
    assert app is not None
    await sync_group_members()
    members = await _group_human_members()
    member_ids = {m.user_id for m in members}
    verified_records = await app.storage.list_verifications_for_user_ids(member_ids)
    verified_ids = {r["user_id"] for r in verified_records}
    return member_ids, verified_ids, members


async def get_live_group_stats() -> Dict[str, int]:
    assert app is not None
    member_ids, verified_ids, _ = await _live_membership_verification()
    all_verified_ids = {int(v["user_id"]) for v in await app.storage.list_all_verifications()}
    table_all_ids, table_paid_ids = await _table_payment_user_ids()

    verified_not_paid = sum(
        1 for uid in verified_ids if uid in table_all_ids and uid not in table_paid_ids
    )
    paid_total = len(table_paid_ids)
    paid_verified = len(table_paid_ids & all_verified_ids)
    paid_unverified = len(table_paid_ids - all_verified_ids)

    return {
        "in_group": len(member_ids),
        "verified": len(verified_ids),
        "unverified": max(0, len(member_ids) - len(verified_ids)),
        "admins": await app.storage.count_admins(),
        "verified_not_paid": verified_not_paid,
        "paid_total": paid_total,
        "paid_verified": paid_verified,
        "paid_unverified": paid_unverified,
    }


async def _group_verified_export_rows() -> list[dict]:
    """Верифицированные участники группы: уникальные user_id (как в статистике)."""
    assert app is not None
    _, verified_ids, _ = await _live_membership_verification()
    if not verified_ids:
        return []

    sqlite_by_id = {
        r["user_id"]: r
        for r in await app.storage.list_verifications_for_user_ids(verified_ids)
    }
    gas_by_id: dict[int, dict] = {}
    for item in await app.gas.list_users():
        user_id = _gas_user_id(item)
        if user_id is None or user_id not in verified_ids or user_id in gas_by_id:
            continue
        gas_by_id[user_id] = item

    rows: list[dict] = []
    for user_id in sorted(verified_ids):
        gas = gas_by_id.get(user_id)
        local = sqlite_by_id.get(user_id)
        if gas:
            rows.append({
                "house": gas.get("house", ""),
                "entrance": gas.get("entrance", ""),
                "apartment": gas.get("apartment", ""),
                "display_name": gas.get("display_name", ""),
                "user_id": gas.get("user_id", ""),
            })
        elif local:
            rows.append({
                "house": local.get("house", ""),
                "entrance": local.get("entrance", ""),
                "apartment": local.get("apartment", ""),
                "display_name": local.get("display_name", ""),
                "user_id": str(user_id),
            })
    return rows


async def get_unverified_action_users() -> List[dict]:
    assert app is not None
    _, verified_ids, members = await _live_membership_verification()
    result: List[dict] = []
    for user in members:
        if user.user_id in verified_ids:
            continue
        result.append({"user_id": user.user_id, "username": user.username, "full_name": user.full_name})

    logger.info("Unverified action users collected: %s", len(result))
    return result


def _parse_user_ids_from_gas_users(users: list[dict]) -> list[int]:
    ids: list[int] = []
    seen: set[int] = set()
    for item in users:
        raw = str(item.get("user_id", "")).strip()
        if not raw.isdigit():
            continue
        user_id = int(raw)
        if user_id in seen:
            continue
        seen.add(user_id)
        ids.append(user_id)
    return ids


async def get_message_recipient_ids(audience: str) -> list[int]:
    """audience: 'verified' — SQLite; 'paid' — строки таблицы с payed=1."""
    assert app is not None
    if audience == "paid":
        users = await app.gas.list_paid_users()
        return _parse_user_ids_from_gas_users(users)
    verifications = await app.storage.list_all_verifications()
    return [int(v["user_id"]) for v in verifications]


async def fail_verification_attempt(message: Message, state: FSMContext, reason: str) -> None:
    """Фиксирует неудачную попытку верификации и сообщает, сколько осталось попыток."""
    assert app is not None
    attempts_left = await app.storage.decrement_attempts(message.from_user.id)
    await state.clear()

    if attempts_left <= 0:
        await message.answer(
            f"{reason}\nВерификация отклонена. Попытки закончились. Обратитесь к администратору.",
            reply_markup=await build_main_menu(message.from_user.id),
        )
    else:
        await message.answer(
            f"{reason}\nОсталось попыток: {attempts_left}.",
            reply_markup=await build_main_menu(message.from_user.id),
        )


async def sync_table_with_group() -> dict:
    """Удаляет из таблицы верификации пользователей, которых уже нет в группе."""
    assert app is not None
    await sync_group_members()

    members = await app.membership.list_all_human_members()
    member_ids = {m.user_id for m in members}

    table_users = await app.gas.list_users()
    removed = 0
    kept = 0
    failed = 0

    for item in table_users:
        try:
            user_id = int(item.get("user_id"))
        except Exception:
            failed += 1
            continue

        if user_id in member_ids:
            kept += 1
            continue

        try:
            if await app.gas.remove_user(user_id):
                await app.storage.delete_verification(user_id)
                removed += 1
            else:
                failed += 1
        except Exception:
            logger.exception("Failed to remove absent user verification: %s", user_id)
            failed += 1

    return {"removed": removed, "kept": kept, "failed": failed}


@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext) -> None:
    assert app is not None
    await state.clear()
    if not is_private_message(message):
        await reject_group_interaction(message)
        return
    await app.storage.touch_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    await message.answer("Привет, я бот, который поможет вам верифицироваться в чате жильцов сообщества домов. Я умею добавлять и удалять верификацию", reply_markup=await build_main_menu(message.from_user.id))


@router.message(F.new_chat_members)
async def handle_new_chat_members(message: Message) -> None:
    """Реакция на вход новых участников в общий чат."""
    assert app is not None

    if is_private_message(message):
        return

    if message.chat.id != app.settings.group_id:
        return

    new_users = list(message.new_chat_members or [])
    human_users = [user for user in new_users if not getattr(user, "is_bot", False)]

    if not human_users:
        return

    for user in human_users:
        await app.storage.touch_user(user.id, user.username, user.full_name)

    mentions = " ".join(user_mention_from_aiogram(user) for user in human_users)
    group_text = f"{mentions}\n\n{new_member_instruction_text()}"

    try:
        async with group_announcement_topic() as thread_id:
            await app.bot.send_message(
                app.settings.group_id,
                group_text,
                message_thread_id=thread_id,
            )
    except Exception:
        logger.exception("Failed to send new member welcome message to group")

    for user in human_users:
        try:
            await app.bot.send_message(
                user.id,
                new_member_instruction_text(),
                reply_markup=await build_main_menu(user.id),
            )
        except Exception:
            logger.info(
                "Could not send private welcome message to user_id=%s. User probably has not started the bot yet.",
                user.id,
            )


@router.message(Command("admin"))
@router.message(F.text == ADMIN_BTN)
async def cmd_admin(message: Message, state: FSMContext) -> None:
    if not is_private_message(message):
        await reject_group_interaction(message, "Админ-панель доступна только в личных сообщениях с ботом.")
        return
    if not await is_admin(message.from_user.id):
        await message.answer("Эта команда доступна только администраторам.")
        return
    await state.clear()
    await message.answer("Админ-панель", reply_markup=admin_menu())


@router.message(F.text == CANCEL_BTN)
async def cancel_action(message: Message, state: FSMContext) -> None:
    if not is_private_message(message):
        return
    await state.clear()
    await message.answer("Действие отменено.", reply_markup=await build_main_menu(message.from_user.id))


@router.message(F.text == VERIFY_BTN)
async def start_verification(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        await reject_group_interaction(message)
        return
    if await app.storage.get_verification(message.from_user.id):
        await message.answer("Вы уже верифицированы.", reply_markup=await build_main_menu(message.from_user.id))
        return
    attempts_left = await app.storage.get_attempts_left(message.from_user.id)
    if attempts_left <= 0:
        await message.answer("Верификация недоступна. Попытки закончились. Обратитесь к администратору.", reply_markup=await build_main_menu(message.from_user.id))
        return
    houses = await app.gas.get_houses()

    await state.set_state(VerifyStates.waiting_for_house)

    await message.answer(
        "Выберите номер дома",
        reply_markup=houses_menu(houses),
    )


@router.message(F.text == DELETE_VERIFY_BTN)
async def delete_verification(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    await state.clear()
    if not await app.storage.get_verification(message.from_user.id):
        await message.answer("У вас нет активной верификации.", reply_markup=await build_main_menu(message.from_user.id))
        return
    try:
        await app.gas.remove_user(message.from_user.id)
    except GasClientError:
        logger.exception("Failed to remove verification from GAS")
        await send_base_error(message)
        return
    await app.storage.delete_verification(message.from_user.id)
    await message.answer("Верификация удалена.", reply_markup=await build_main_menu(message.from_user.id))


@router.message(StateFilter(VerifyStates.waiting_for_house))
async def process_house(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    house_input = (message.text or "").strip()
    try:
        house = await app.gas.resolve_house(house_input)
    except GasClientError:
        logger.exception("GasClientError while checking house")
        await send_base_error(message)
        return
    if house is None:
        if not valid_house_input(house_input.lower()):
            await fail_verification_attempt(
                message, state, "Авторизация отклонена. Номер дома нужно вводить кириллицей в нижнем регистре."
            )
            return
        await fail_verification_attempt(message, state, "Авторизация отклонена. Такой номер дома не найден.")
        return
    await state.update_data(house=house)
    await state.set_state(VerifyStates.waiting_for_entrance)
    await message.answer("Введите номер подъезда.", reply_markup=ReplyKeyboardRemove())


@router.message(StateFilter(VerifyStates.waiting_for_entrance))
async def process_entrance(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    entrance = (message.text or "").strip()
    data = await state.get_data()
    try:
        exists = await app.gas.check_entrance(data["house"], entrance)
    except GasClientError:
        logger.exception("GasClientError while checking entrance")
        await send_base_error(message)
        return
    if not exists:
        await fail_verification_attempt(message, state, "Авторизация отклонена. Такой подъезд не найден для указанного дома.")
        return
    await state.update_data(entrance=entrance)
    await state.set_state(VerifyStates.waiting_for_floor)
    await message.answer("Введите этаж.")


@router.message(StateFilter(VerifyStates.waiting_for_floor))
async def process_floor(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    floor = (message.text or "").strip()
    data = await state.get_data()
    try:
        exists = await app.gas.check_floor(data["house"], data["entrance"], floor)
    except GasClientError:
        logger.exception("GasClientError while checking floor")
        await send_base_error(message)
        return
    if not exists:
        await fail_verification_attempt(message, state, "Авторизация отклонена. Такой этаж не найден для указанного дома и подъезда.")
        return
    await state.update_data(floor=floor)
    await state.set_state(VerifyStates.waiting_for_apartment)
    await message.answer("Введите номер квартиры.")


@router.message(StateFilter(VerifyStates.waiting_for_apartment))
async def process_apartment(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    apartment = (message.text or "").strip()
    data = await state.get_data()
    try:
        exists = await app.gas.check_apartment(data["house"], data["entrance"], data["floor"], apartment)
        if not exists:
            await fail_verification_attempt(message, state, "Авторизация отклонена. Такая квартира не найдена.")
            return
        display_name = message_user_display_name(message)
        await app.gas.verify_user(
            house=data["house"], entrance=data["entrance"], floor=data["floor"], apartment=apartment,
            user_id=message.from_user.id, display_name=display_name, username=message.from_user.username or "",
        )
        await app.storage.save_verification(
            user_id=message.from_user.id, username=message.from_user.username, full_name=message.from_user.full_name,
            display_name=display_name, house=data["house"], entrance=data["entrance"], floor=data["floor"], apartment=apartment,
        )
        await app.storage.touch_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
        app.gas.invalidate_cache()

        # Важно: после успешной верификации автоматически снимаем ограничение отправки.
        try:
            await app.membership.unrestrict_user_sending(message.from_user.id)
        except Exception:
            logger.exception("Failed to auto-unrestrict user after verification: %s", message.from_user.id)

        await state.clear()
        await message.answer(
            f"Вы верифицированы: дом {html.escape(data['house'])}, подъезд {html.escape(data['entrance'])}, этаж {html.escape(data['floor'])}, квартира {html.escape(apartment)}.",
            reply_markup=await build_main_menu(message.from_user.id),
        )
    except GasClientError as exc:
        logger.exception("Verification failed")
        if "Превышено количество зарегистрированных пользователей на квартиру" in str(exc):
            await send_verification_error(message, "Превышено количество зарегистрированных пользователей на квартиру")
        else:
            await send_verification_error(message)


@router.callback_query(F.data == "admin:stats")
async def admin_stats(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    stats = await get_live_group_stats()
    await callback.message.answer(
        "<b>Статистика</b>\n"
        f"Всего в группе: {stats['in_group']}\n"
        f"Верифицировано: {stats['verified']}\n"
        f"Неверифицировано: {stats['unverified']}\n"
        f"Администраторов: {stats['admins']}\n"
        f"Верифицированные не оплатившие: {stats['verified_not_paid']}\n"
        f"Оплативших: {stats['paid_total']}\n"
        f"Оплатившие верифицированные: {stats['paid_verified']}\n"
        f"Оплатившие не верифицированные: {stats['paid_unverified']}"
    )
    await callback.answer()


@router.callback_query(F.data == "admin:unverified")
async def admin_unverified(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await get_unverified_action_users()
    if not users:
        await callback.message.answer("Неверифицированных пользователей сейчас нет.")
        await callback.answer()
        return
    lines = ["<b>Неверифицированные пользователи</b>"]
    for user in users[:100]:
        lines.append(f"• {await mention_html_live(user['user_id'])}")
    await callback.message.answer("\n".join(lines))
    await callback.answer()


@router.callback_query(F.data == "admin:unverified_actions")
async def admin_unverified_actions(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("Действия с неверифицированными", reply_markup=unverified_actions_menu())
    await callback.answer()


@router.callback_query(F.data == "admin:verified_actions")
async def admin_verified_actions(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("Верифицированные", reply_markup=verified_actions_menu())
    await callback.answer()


@router.callback_query(F.data == "admin:paid_actions")
async def admin_paid_actions(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(
        "Оплатившие (квартиры с payed=1 в таблице)",
        reply_markup=paid_actions_menu(),
    )
    await callback.answer()


@router.callback_query(F.data == "admin:back")
async def admin_back(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("Админ-панель", reply_markup=admin_menu())
    await callback.answer()


async def _open_group_topic(chat_id: int, thread_id: int | None) -> None:
    assert app is not None
    if not thread_id:
        return
    try:
        if thread_id == 1:
            await app.bot.reopen_general_forum_topic(chat_id=chat_id)
        else:
            await app.bot.reopen_forum_topic(chat_id=chat_id, message_thread_id=thread_id)
    except Exception:
        logger.warning(
            "Failed to reopen forum topic %s in chat %s",
            thread_id,
            chat_id,
            exc_info=True,
        )


async def _close_group_topic(chat_id: int, thread_id: int | None) -> None:
    assert app is not None
    if not thread_id:
        return
    try:
        if thread_id == 1:
            await app.bot.close_general_forum_topic(chat_id=chat_id)
        else:
            await app.bot.close_forum_topic(chat_id=chat_id, message_thread_id=thread_id)
    except Exception:
        logger.warning(
            "Failed to close forum topic %s in chat %s",
            thread_id,
            chat_id,
            exc_info=True,
        )


@asynccontextmanager
async def group_announcement_topic():
    assert app is not None
    chat_id = app.settings.group_id
    thread_id = app.settings.announcement_thread_id
    await _open_group_topic(chat_id, thread_id)
    try:
        yield thread_id
    finally:
        await _close_group_topic(chat_id, thread_id)


async def send_group_mentions(text: str, users: List[dict]) -> int:
    assert app is not None
    if not users:
        return 0
    mentions = []
    for u in users:
        mentions.append(await mention_html_live(u["user_id"]))
    chunks: List[str] = []
    current = ""
    for mention in mentions:
        candidate = (current + " " + mention).strip()
        if len(candidate) > 3000 and current:
            chunks.append(current)
            current = mention
        else:
            current = candidate
    if current:
        chunks.append(current)
    sent = 0
    async with group_announcement_topic() as thread_id:
        for mention_chunk in chunks:
            await app.bot.send_message(
                app.settings.group_id,
                f"{text}\n\n{mention_chunk}",
                message_thread_id=thread_id,
            )
            sent += 1
    return sent


@router.callback_query(F.data == "admin:warn")
async def admin_warn(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await get_unverified_action_users()
    if not users:
        await callback.message.answer("Неверифицированных пользователей сейчас нет.")
    else:
        sent = await send_group_mentions("Вы не прошли верификацию. Пожалуйста, напишите боту в личные сообщения и пройдите проверку.", users)
        await callback.message.answer(f"Предупреждение отправлено в группу. Сообщений: {sent}.")
    await callback.answer()


@router.callback_query(F.data == "admin:restrict_unverified")
async def admin_restrict_unverified(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await get_unverified_action_users()
    if not users:
        await callback.message.answer("Неверифицированных пользователей сейчас нет.")
        await callback.answer()
        return
    restricted = 0
    failed = 0
    for user in users:
        if await app.membership.restrict_user_sending(user["user_id"]):
            restricted += 1
        else:
            failed += 1
    try:
        await send_group_mentions("Ограничена отправка сообщений не верифицированным пользователям, для возобновления возможности отправки необходимо верифицироваться", users)
    except Exception:
        logger.exception("Failed to notify group about restricted unverified users")
    await callback.message.answer(f"Ограничение применено. Ограничено: {restricted}. Не удалось ограничить: {failed}.")
    await callback.answer()


@router.callback_query(F.data == "admin:unrestrict_unverified")
async def admin_unrestrict_unverified(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await get_unverified_action_users()
    if not users:
        await callback.message.answer("Неверифицированных пользователей сейчас нет.")
        await callback.answer()
        return
    unrestricted = 0
    failed = 0
    for user in users:
        if await app.membership.unrestrict_user_sending(user["user_id"]):
            unrestricted += 1
        else:
            failed += 1
    await callback.message.answer(f"Ограничение отправки отменено. Снято ограничений: {unrestricted}. Не удалось снять: {failed}.")
    await callback.answer()


@router.callback_query(F.data == "admin:purge")
async def admin_purge(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await get_unverified_action_users()
    removed = 0
    failed = 0
    for user in users:
        if await app.membership.remove_user(user["user_id"]):
            removed += 1
        else:
            failed += 1
    await callback.message.answer(f"Удаление завершено. Удалено пользователей: {removed}. Не удалось удалить: {failed}.")
    await callback.answer()


@router.callback_query(F.data == "admin:revoke_verification")
async def admin_revoke_verification(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_revoke_user_id)
    await callback.message.answer("Введите идентификатор пользователя, с которого нужно снять верификацию.")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_revoke_user_id))
async def admin_revoke_find_user(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Введите числовой идентификатор пользователя.")
        return
    user_id = int(text)
    try:
        user = await app.gas.find_user(user_id)
    except GasClientError:
        logger.exception("Failed to find user in GAS before revoke")
        await send_base_error(message)
        return
    if not user:
        await state.clear()
        await message.answer("Пользователь с таким идентификатором не найден в таблице.", reply_markup=await build_main_menu(message.from_user.id))
        return
    await state.update_data(revoke_user_id=user_id)
    house = html.escape(str(user.get("house", "")))
    entrance = html.escape(str(user.get("entrance", "")))
    apartment = html.escape(str(user.get("apartment", "")))
    await message.answer(
        f"Снять верификацию с пользователя {house}, {entrance}, {apartment}?",
        reply_markup=confirm_revoke_menu(user_id),
    )


@router.callback_query(F.data.startswith("admin:revoke_confirm:"))
async def admin_revoke_confirm(callback: CallbackQuery, state: FSMContext) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    user_id = int(callback.data.split(":")[-1])
    try:
        found = await app.gas.find_user(user_id)
        if not found:
            await callback.message.answer("Пользователь с таким идентификатором уже не найден в таблице.")
            await state.clear()
            await callback.answer()
            return
        await app.gas.remove_user(user_id)
    except GasClientError:
        logger.exception("Failed to revoke verification in GAS")
        await callback.message.answer("Не удалось снять верификацию в таблице.")
        await callback.answer()
        return
    await app.storage.delete_verification(user_id)
    await state.clear()
    await callback.message.answer(f"Верификация пользователя {user_id} снята.")
    await callback.answer()


@router.callback_query(F.data == "admin:revoke_cancel")
async def admin_revoke_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await callback.message.answer("Снятие верификации отменено.")
    await callback.answer()



@router.callback_query(F.data == "admin:reset_attempts")
async def admin_reset_attempts(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_reset_attempts_user_id)
    await callback.message.answer("Введите идентификатор пользователя, которому нужно обнулить счетчик верификаций.")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_reset_attempts_user_id))
async def admin_reset_attempts_apply(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return

    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Введите числовой идентификатор пользователя.")
        return

    user_id = int(text)

    local_verified = await app.storage.get_verification(user_id)
    try:
        table_verified = await app.gas.find_user(user_id)
    except GasClientError:
        logger.exception("Failed to check user in GAS before resetting attempts")
        await send_base_error(message)
        return

    if local_verified or table_verified:
        await state.clear()
        await message.answer("Пользователь уже верифицирован. Счетчик можно обнулить только для неверифицированного пользователя.", reply_markup=await build_main_menu(message.from_user.id))
        return

    await app.storage.set_attempts_left(user_id, 2)
    await state.clear()
    await message.answer(f"Счетчик верификаций для пользователя {user_id} обнулен. Доступно попыток: 2.", reply_markup=await build_main_menu(message.from_user.id))

    try:
        await app.bot.send_message(user_id, "Администратор обнулил ваш счетчик попыток верификации. Вам снова доступно 2 попытки.")
    except Exception:
        logger.exception("Failed to notify user about attempts reset: %s", user_id)


@router.callback_query(F.data == "admin:sync_table")
async def admin_sync_table(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        result = await sync_table_with_group()
    except GasClientError:
        logger.exception("Failed to sync table with group")
        await callback.message.answer("Не удалось синхронизировать таблицу с группой.")
        await callback.answer()
        return

    await callback.message.answer(
        "Синхронизация завершена.\n"
        f"Оставлено в таблице: {result['kept']}\n"
        f"Удалено из таблицы: {result['removed']}\n"
        f"Ошибок: {result['failed']}"
    )
    await callback.answer()


@router.callback_query(F.data == "admin:list_admins")
async def admin_list_admins(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    admins = await app.storage.list_admins()
    lines = ["<b>Администраторы</b>"]
    for item in admins:
        lines.append(f"• id={item['user_id']}")
    await callback.message.answer("\n".join(lines))
    await callback.answer()


@router.callback_query(F.data == "admin:add_admin")
async def admin_add_admin(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_add_admin)
    await callback.message.answer("Пришлите числовой user_id нового администратора.")
    await callback.answer()


@router.callback_query(F.data == "admin:remove_admin")
async def admin_remove_admin(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_remove_admin)
    await callback.message.answer("Пришлите числовой user_id администратора, которого нужно удалить.")
    await callback.answer()


def _extract_user_id(message: Message) -> int | None:
    text = (message.text or "").strip()
    return int(text) if text.isdigit() else None


@router.message(StateFilter(AdminStates.waiting_for_add_admin))
async def do_add_admin(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    user_id = _extract_user_id(message)
    if not user_id:
        await message.answer("Пришлите числовой user_id.")
        return
    await app.storage.add_admin(user_id, message.from_user.id)
    await state.clear()
    await message.answer(f"user_id={user_id} добавлен в администраторы.", reply_markup=await build_main_menu(message.from_user.id))


@router.message(StateFilter(AdminStates.waiting_for_remove_admin))
async def do_remove_admin(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    user_id = _extract_user_id(message)
    if not user_id:
        await message.answer("Пришлите числовой user_id.")
        return
    admins = await app.storage.list_admins()
    if len(admins) <= 1:
        await message.answer("Нельзя удалить последнего администратора.")
        return
    await app.storage.remove_admin(user_id)
    await state.clear()
    await message.answer(f"user_id={user_id} удалён из администраторов.", reply_markup=await build_main_menu(message.from_user.id))



@router.callback_query(F.data == "admin:export_verified")
async def admin_export_verified(callback: CallbackQuery) -> None:
    assert app is not None

    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return

    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    users = await _group_verified_export_rows()

    if not users:
        await callback.message.answer("Нет верифицированных участников группы.")
        await callback.answer()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Верифицированные"

    ws.append(["Дом", "Подъезд", "Квартира", "Имя пользователя", "Идентификатор"])

    for user in users:
        ws.append([
            user.get("house", ""),
            user.get("entrance", ""),
            user.get("apartment", ""),
            user.get("display_name", ""),
            user.get("user_id", ""),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    document = BufferedInputFile(buffer.read(), filename="verified_users.xlsx")

    await callback.message.answer(f"Верифицированных пользователей: {len(users)}")
    await callback.message.answer_document(document)
    await callback.answer()


@router.callback_query(F.data == "paid:export")
async def admin_export_paid(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        users = await app.gas.list_paid_users()
    except GasClientError:
        logger.exception("Failed to export paid users")
        await callback.message.answer("Не удалось загрузить список из таблицы.")
        await callback.answer()
        return
    if not users:
        await callback.message.answer("Нет пользователей в квартирах с payed=1.")
        await callback.answer()
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Оплатившие"
    ws.append(["Дом", "Подъезд", "Квартира", "Имя пользователя", "Идентификатор", "payed"])
    for user in users:
        ws.append([
            user.get("house", ""),
            user.get("entrance", ""),
            user.get("apartment", ""),
            user.get("display_name", ""),
            user.get("user_id", ""),
            user.get("payed", "1"),
        ])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    document = BufferedInputFile(buffer.read(), filename="paid_users.xlsx")
    await callback.message.answer(f"Оплативших пользователей: {len(users)}")
    await callback.message.answer_document(document)
    await callback.answer()


@router.callback_query(F.data == "verified:remove_limit")
async def admin_remove_limit_start(callback: CallbackQuery, state: FSMContext) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        houses = await app.gas.get_houses()
    except GasClientError:
        logger.exception("Failed to fetch houses for remove_limit")
        await callback.message.answer("Ошибка загрузки списка домов. Попробуйте позже.")
        await callback.answer()
        return
    if not houses:
        await callback.message.answer("Список домов пуст.")
        await callback.answer()
        return
    await state.set_state(AdminStates.waiting_for_remove_limit_house)
    await callback.message.answer("Выберите дом:", reply_markup=houses_menu(houses))
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_remove_limit_house))
async def admin_remove_limit_house(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    house_input = (message.text or "").strip()
    if not house_input:
        await message.answer("Пожалуйста, выберите дом из списка.")
        return
    try:
        house = await app.gas.resolve_house(house_input)
    except GasClientError:
        logger.exception("GasClientError while checking house in remove_limit")
        await send_base_error(message)
        return
    if house is None:
        await message.answer("Такой дом не найден. Выберите из предложенных вариантов.")
        return
    await state.update_data(remove_limit_house=house)
    await state.set_state(AdminStates.waiting_for_remove_limit_apartment)
    await message.answer(f"Дом {house} выбран. Введите номер квартиры:", reply_markup=ReplyKeyboardRemove())


@router.message(StateFilter(AdminStates.waiting_for_remove_limit_apartment))
async def admin_remove_limit_apartment(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    apartment = (message.text or "").strip()
    if not apartment:
        await message.answer("Введите номер квартиры.")
        return
    data = await state.get_data()
    house = data["remove_limit_house"]
    try:
        if not await app.gas.apartment_exists_for_house(house, apartment):
            await message.answer(
                f"Квартира {html.escape(apartment)} в доме {html.escape(house)} не найдена в таблице.",
                reply_markup=await build_main_menu(message.from_user.id),
            )
            await state.clear()
            return
        await app.gas.set_apartment_limit(house, apartment, 10)
    except GasClientError:
        logger.exception("Failed to set apartment limit house=%s apartment=%s", house, apartment)
        await send_base_error(message)
        return
    await state.clear()
    await message.answer(
        f"Лимит для квартиры {html.escape(apartment)} (дом {html.escape(house)}) увеличен до 10 пользователей.",
        reply_markup=await build_main_menu(message.from_user.id),
    )



@router.callback_query(F.data == "admin:manual_verify")
async def admin_manual_verify_start(callback: CallbackQuery, state: FSMContext) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_manual_verify_user_id)
    await callback.message.answer("Введите Telegram ID пользователя:")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_manual_verify_user_id))
async def admin_manual_verify_user_id(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    raw = (message.text or "").strip()
    if not raw.lstrip("-").isdigit():
        await message.answer("ID должен быть числом. Попробуйте ещё раз:")
        return
    user_id = int(raw)
    await state.update_data(manual_verify_user_id=user_id)
    await state.set_state(AdminStates.waiting_for_manual_verify_house)
    try:
        houses = await app.gas.get_houses()
    except GasClientError:
        await send_base_error(message)
        return
    await message.answer("Выберите дом:", reply_markup=houses_menu(houses))


@router.message(StateFilter(AdminStates.waiting_for_manual_verify_house))
async def admin_manual_verify_house(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    house_input = (message.text or "").strip()
    try:
        house = await app.gas.resolve_house(house_input)
    except GasClientError:
        await send_base_error(message)
        return
    if house is None:
        await message.answer("Такой дом не найден. Выберите из списка.")
        return
    await state.update_data(manual_verify_house=house)
    await state.set_state(AdminStates.waiting_for_manual_verify_apartment)
    await message.answer(f"Дом {html.escape(house)} выбран. Введите номер квартиры:", reply_markup=ReplyKeyboardRemove())


@router.message(StateFilter(AdminStates.waiting_for_manual_verify_apartment))
async def admin_manual_verify_apartment(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    apartment = (message.text or "").strip()
    if not apartment:
        await message.answer("Введите номер квартиры.")
        return
    data = await state.get_data()
    house = data["manual_verify_house"]
    user_id = data["manual_verify_user_id"]

    # Find any row matching house+apartment (ignore entrance/floor)
    try:
        rows = await app.gas._get_rows()
    except GasClientError:
        await send_base_error(message)
        return

    matched = next(
        (r for r in rows if r["house"].lower() == house.lower() and r["apartment"] == apartment),
        None
    )
    if not matched:
        await state.clear()
        await message.answer(
            f"Квартира {html.escape(apartment)} в доме {html.escape(house)} не найдена в таблице.",
            reply_markup=await build_main_menu(message.from_user.id),
        )
        return

    entrance = matched.get("entrance", "")
    floor = matched.get("floor", "")

    # Fetch real Telegram name
    username, full_name = await get_live_display_name(user_id)
    display_name = full_name
    try:
        await app.gas.verify_user(house, entrance, floor, apartment, user_id, display_name, username or "")
        await app.storage.save_verification(
            user_id=user_id,
            username=username,
            full_name=full_name,
            display_name=display_name,
            house=house,
            entrance=entrance,
            floor=floor,
            apartment=apartment,
        )
        app.gas.invalidate_cache()
        try:
            await app.membership.unrestrict_user_sending(user_id)
        except Exception:
            logger.exception("Failed to unrestrict user %s after manual verify", user_id)
    except GasClientError as e:
        await state.clear()
        await message.answer(f"Ошибка верификации: {html.escape(str(e))}", reply_markup=await build_main_menu(message.from_user.id))
        return

    await state.clear()
    await message.answer(
        f"Пользователь {user_id} верифицирован: дом {html.escape(house)}, квартира {html.escape(apartment)}.",
        reply_markup=await build_main_menu(message.from_user.id),
    )
    try:
        await app.bot.send_message(
            user_id,
            f"✅ Вы верифицированы администратором.\n"
            f"Дом: <b>{html.escape(house)}</b>, квартира: <b>{html.escape(apartment)}</b>.\n"
            f"Теперь вы можете участвовать в чате.",
        )
    except Exception:
        logger.warning("Could not notify user %s about manual verification", user_id)


# ── Broadcast ──────────────────────────────────────────────────

@router.callback_query(F.data == "admin:broadcast")
async def admin_broadcast_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.update_data(message_audience="verified")
    await state.set_state(AdminStates.waiting_for_broadcast_message)
    await callback.message.answer(
        "Введите текст сообщения для рассылки всем верифицированным.\n"
        "Поддерживается HTML-форматирование (<b>жирный</b>, <i>курсив</i>).\n"
        "Для отмены нажмите Отмена."
    )
    await callback.answer()


@router.callback_query(F.data == "paid:broadcast")
async def admin_paid_broadcast_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.update_data(message_audience="paid")
    await state.set_state(AdminStates.waiting_for_broadcast_message)
    await callback.message.answer(
        "Введите текст сообщения для рассылки оплатившим (квартиры с payed=1 в таблице).\n"
        "Поддерживается HTML-форматирование (<b>жирный</b>, <i>курсив</i>).\n"
        "Для отмены нажмите Отмена."
    )
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_broadcast_message))
async def admin_broadcast_send(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    if (message.text or "").strip() == CANCEL_BTN:
        await state.clear()
        await message.answer("Отменено.", reply_markup=await build_main_menu(message.from_user.id))
        return

    text = message.text or message.caption or ""
    if not text.strip():
        await message.answer("Сообщение не может быть пустым.")
        return

    data = await state.get_data()
    audience = data.get("message_audience", "verified")
    try:
        recipient_ids = await get_message_recipient_ids(audience)
    except GasClientError:
        logger.exception("Failed to load recipients for broadcast audience=%s", audience)
        await send_base_error(message)
        return

    if not recipient_ids:
        await state.clear()
        label = "оплативших" if audience == "paid" else "верифицированных"
        await message.answer(f"Нет {label} пользователей.", reply_markup=await build_main_menu(message.from_user.id))
        return

    await state.clear()
    sent, failed = 0, 0
    progress_msg = await message.answer(f"Отправляю... 0 / {len(recipient_ids)}")

    for i, user_id in enumerate(recipient_ids):
        try:
            await app.bot.send_message(user_id, text)
            sent += 1
        except Exception:
            failed += 1
        if (i + 1) % 10 == 0:
            try:
                await progress_msg.edit_text(f"Отправляю... {i+1} / {len(recipient_ids)}")
            except Exception:
                pass

    await progress_msg.edit_text(
        f"Рассылка завершена.\nОтправлено: {sent}, не доставлено: {failed}."
    )
    await message.answer("Готово.", reply_markup=await build_main_menu(message.from_user.id))


# ── Create Poll ─────────────────────────────────────────────────

@router.callback_query(F.data == "admin:create_poll")
async def admin_create_poll_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.update_data(message_audience="verified")
    await state.set_state(AdminStates.waiting_for_poll_question)
    await callback.message.answer("Введите вопрос голосования (для верифицированных):")
    await callback.answer()


@router.callback_query(F.data == "paid:create_poll")
async def admin_paid_create_poll_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.update_data(message_audience="paid")
    await state.set_state(AdminStates.waiting_for_poll_question)
    await callback.message.answer("Введите вопрос голосования (для оплативших):")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_poll_question))
async def admin_poll_question(message: Message, state: FSMContext) -> None:
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    question = (message.text or "").strip()
    if not question:
        await message.answer("Вопрос не может быть пустым.")
        return
    await state.update_data(poll_question=question, poll_options=[])
    await state.set_state(AdminStates.waiting_for_poll_options)
    await message.answer(
        f"Вопрос: <b>{html.escape(question)}</b>\n\n"
        "Теперь вводите варианты ответов по одному.\n"
        "Когда добавите все варианты (минимум 2), напишите <b>Готово</b>."
    )


@router.message(StateFilter(AdminStates.waiting_for_poll_options))
async def admin_poll_options(message: Message, state: FSMContext) -> None:
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    data = await state.get_data()
    options = data.get("poll_options", [])

    if text.lower() == "готово":
        if len(options) < 2:
            await message.answer("Нужно минимум 2 варианта.")
            return
        # Create poll immediately, no deadline
        assert app is not None
        poll_id = await app.storage.create_poll(
            question=data["poll_question"],
            options=options,
            created_by=message.from_user.id,
            closes_at=None,
        )
        await state.clear()

        poll_options = await app.storage.get_poll_options(poll_id)
        audience = data.get("message_audience", "verified")
        try:
            recipient_ids = await get_message_recipient_ids(audience)
        except GasClientError:
            logger.exception("Failed to load recipients for poll audience=%s", audience)
            await send_base_error(message)
            return
        if not recipient_ids:
            label = "оплативших" if audience == "paid" else "верифицированных"
            await message.answer(f"Нет {label} пользователей для рассылки голосования.")
            return

        poll_text = (
            f"📊 <b>Голосование #{poll_id}</b>\n\n"
            f"{html.escape(data['poll_question'])}\n\n"
            f"Выберите вариант ответа:"
        )

        sent, failed = 0, 0
        progress_msg = await message.answer(f"Рассылаю голосование... 0 / {len(recipient_ids)}")

        for i, user_id in enumerate(recipient_ids):
            try:
                await app.bot.send_message(
                    user_id,
                    poll_text,
                    reply_markup=poll_vote_keyboard(poll_id, poll_options),
                )
                sent += 1
            except Exception:
                failed += 1
            if (i + 1) % 10 == 0:
                try:
                    await progress_msg.edit_text(f"Рассылаю голосование... {i+1} / {len(recipient_ids)}")
                except Exception:
                    pass

        await progress_msg.edit_text(
            f"Голосование #{poll_id} разослано.\nДоставлено: {sent}, не доставлено: {failed}."
        )
        await message.answer(
            f"Управление голосованием #{poll_id}:",
            reply_markup=poll_manage_keyboard(poll_id),
        )
        return

    if len(options) >= 10:
        await message.answer("Максимум 10 вариантов. Напишите Готово для продолжения.")
        return

    options.append(text)
    await state.update_data(poll_options=options)
    await message.answer(
        f"Вариант {len(options)} добавлен: <b>{html.escape(text)}</b>\n"
        f"Всего вариантов: {len(options)}. Добавьте ещё или напишите <b>Готово</b>."
    )



# ── Vote handler ────────────────────────────────────────────────

@router.callback_query(F.data.startswith("vote:"))
async def handle_vote(callback: CallbackQuery) -> None:
    assert app is not None
    _, poll_id_str, option_id_str = callback.data.split(":")
    poll_id, option_id = int(poll_id_str), int(option_id_str)
    user_id = callback.from_user.id

    # Only verified users can vote
    verification = await app.storage.get_verification(user_id)
    if not verification:
        await callback.answer("Голосовать могут только верифицированные пользователи.", show_alert=True)
        return

    poll = await app.storage.get_poll(poll_id)
    if not poll:
        await callback.answer("Голосование не найдено.", show_alert=True)
        return
    if poll["is_closed"]:
        await callback.answer("Это голосование уже закрыто.", show_alert=True)
        return

    existing = await app.storage.get_user_vote(poll_id, user_id)
    await app.storage.save_vote(poll_id, user_id, option_id)

    options = await app.storage.get_poll_options(poll_id)
    chosen = next((o for o in options if o["option_id"] == option_id), None)
    chosen_text = chosen["option_text"] if chosen else "?"

    if existing:
        await callback.answer(f"Ваш голос изменён на: {chosen_text}", show_alert=False)
    else:
        await callback.answer(f"Ваш голос: {chosen_text}", show_alert=False)


# ── Poll results & close ────────────────────────────────────────

@router.callback_query(F.data.startswith("poll:results:"))
async def poll_results(callback: CallbackQuery) -> None:
    assert app is not None
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    poll_id = int(callback.data.split(":")[2])
    poll = await app.storage.get_poll(poll_id)
    if not poll:
        await callback.answer("Голосование не найдено.", show_alert=True)
        return

    results = await app.storage.get_poll_results(poll_id)
    total = sum(r["vote_count"] for r in results)
    lines = [f"📊 <b>Результаты голосования #{poll_id}</b>", f"<i>{html.escape(poll['question'])}</i>", ""]
    for r in results:
        pct = round(r["vote_count"] / total * 100) if total > 0 else 0
        bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
        lines.append(f"{html.escape(r['option_text'])}: {r['vote_count']} ({pct}%)")
        lines.append(f"{bar}")
    lines.append(f"\nВсего проголосовало: {total}")
    if poll["is_closed"]:
        lines.append("🔒 Голосование закрыто")

    await callback.message.answer("\n".join(lines))
    await callback.answer()


@router.callback_query(F.data.startswith("poll:close:"))
async def poll_close(callback: CallbackQuery) -> None:
    assert app is not None
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    poll_id = int(callback.data.split(":")[2])
    await app.storage.close_poll(poll_id)

    results = await app.storage.get_poll_results(poll_id)
    poll = await app.storage.get_poll(poll_id)
    total = sum(r["vote_count"] for r in results)

    lines = [f"🔒 <b>Голосование #{poll_id} закрыто</b>", f"<i>{html.escape(poll['question'])}</i>", ""]
    for r in results:
        pct = round(r["vote_count"] / total * 100) if total > 0 else 0
        lines.append(f"{html.escape(r['option_text'])}: {r['vote_count']} ({pct}%)")
    lines.append(f"\nВсего проголосовало: {total}")

    # Notify all voters of final results
    voters = await app.storage.get_poll_voters(poll_id)
    result_text = "\n".join(lines)
    for voter_id in voters:
        try:
            await app.bot.send_message(voter_id, result_text)
        except Exception:
            pass

    await callback.message.answer(result_text)
    await callback.answer("Голосование закрыто.")


# ── Admin: Docs ─────────────────────────────────────────────────

DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "documents")


@router.callback_query(F.data == "admin:docs")
async def admin_docs_menu(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    docs = await app.storage.list_documents()
    text = f"📁 <b>Документация</b>\nВсего документов: {len(docs)}\n\nНажмите на документ чтобы удалить его."
    await callback.message.answer(text, reply_markup=docs_keyboard(docs))
    await callback.answer()


@router.callback_query(F.data == "docs:add")
async def admin_docs_add_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_doc_name)
    await callback.message.answer("Введите наименование документа:")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_doc_name))
async def admin_docs_name(message: Message, state: FSMContext) -> None:
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("Наименование не может быть пустым.")
        return
    await state.update_data(doc_name=name)
    await state.set_state(AdminStates.waiting_for_doc_file)
    await message.answer(f"Наименование: <b>{html.escape(name)}</b>\nТеперь отправьте файл документа.")


@router.message(StateFilter(AdminStates.waiting_for_doc_file))
async def admin_docs_file(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    if not message.document:
        await message.answer("Пожалуйста, отправьте файл.")
        return
    data = await state.get_data()
    name = data["doc_name"]
    doc = message.document
    original_name = doc.file_name or f"{doc.file_id}"
    os.makedirs(DOCS_DIR, exist_ok=True)
    file_path = os.path.join(DOCS_DIR, original_name)
    file = await app.bot.get_file(doc.file_id)
    await app.bot.download_file(file.file_path, destination=file_path)
    await app.storage.add_document(name=name, filename=original_name)
    await state.clear()
    await message.answer(
        f"✅ Документ <b>{html.escape(name)}</b> сохранён.",
        reply_markup=await build_main_menu(message.from_user.id),
    )


@router.callback_query(F.data.startswith("docs:delete:"))
async def admin_docs_delete(callback: CallbackQuery) -> None:
    assert app is not None
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    doc_id = int(callback.data.split(":")[2])
    docs = await app.storage.list_documents()
    doc = next((d for d in docs if d["doc_id"] == doc_id), None)
    if doc:
        file_path = os.path.join(DOCS_DIR, doc["filename"])
        if os.path.exists(file_path):
            os.remove(file_path)
        await app.storage.delete_document(doc_id)
        await callback.answer("Документ удалён.")
    else:
        await callback.answer("Документ не найден.")
    docs = await app.storage.list_documents()
    await callback.message.edit_reply_markup(reply_markup=docs_keyboard(docs))


# ── User: Docs ──────────────────────────────────────────────────

@router.message(F.text == DOCS_BTN)
async def user_docs_menu(message: Message) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    verification = await app.storage.get_verification(message.from_user.id)
    if not verification:
        await message.answer("Документация доступна только верифицированным пользователям.")
        return
    docs = await app.storage.list_documents()
    if not docs:
        await message.answer("Документов пока нет.")
        return
    await message.answer("📁 <b>Документация</b>\nВыберите документ:", reply_markup=user_docs_keyboard(docs))


@router.callback_query(F.data.startswith("docs:get:"))
async def user_docs_get(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    verification = await app.storage.get_verification(callback.from_user.id)
    if not verification:
        await callback.answer("Только для верифицированных.", show_alert=True)
        return
    doc_id = int(callback.data.split(":")[2])
    docs = await app.storage.list_documents()
    doc = next((d for d in docs if d["doc_id"] == doc_id), None)
    if not doc:
        await callback.answer("Документ не найден.", show_alert=True)
        return
    file_path = os.path.join(DOCS_DIR, doc["filename"])
    if not os.path.exists(file_path):
        await callback.answer("Файл не найден на сервере.", show_alert=True)
        return
    await callback.message.answer_document(
        FSInputFile(file_path, filename=doc["filename"]),
        caption=doc["name"],
    )
    await callback.answer()


# ── Admin: Barriers ─────────────────────────────────────────────

async def _open_barrier_and_report(
    target: Message,
    state: FSMContext,
    barrier_key: str,
    pin: str | None = None,
) -> None:
    assert app is not None
    try:
        response = await app.barrier.open_barrier(barrier_key, pin=pin)
    except ChangeDeviceRequired:
        # Запоминаем выбранный шлагбаум и ждём PIN из SMS
        logger.info("Требуется PIN для шлагбаума '%s', ожидаем ввод от пользователя", barrier_key)
        await state.set_state(AdminStates.waiting_for_barrier_pin)
        await state.update_data(barrier_pin_key=barrier_key)
        await target.answer("Введите PIN из SMS")
        return
    except BarrierClientError as exc:
        logger.warning("Не удалось открыть шлагбаум '%s': %s", barrier_key, exc)
        await target.answer(f"❌ Ошибка: {html.escape(str(exc))}")
        return

    if response.ok:
        logger.info("Шлагбаум '%s' успешно открыт", barrier_key)
        await target.answer("✅ Открыт")
    else:
        logger.info("Сервер вернул ошибку для '%s': %s", barrier_key, response.message)
        await target.answer(f"❌ Ошибка: {html.escape(response.message or 'неизвестная ошибка')}")


@router.callback_query(F.data == "admin:barriers")
async def admin_barriers_menu(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(
        "🚧 <b>Шлагбаумы</b>\nВыберите шлагбаум для открытия:",
        reply_markup=barriers_menu(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("barrier:open:"))
async def admin_barrier_open(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    barrier_key = callback.data.removeprefix("barrier:open:")
    barrier = BARRIERS.get(barrier_key)
    if not barrier:
        await callback.answer("Неизвестный шлагбаум", show_alert=True)
        return

    logger.info(
        "Админ %s нажал кнопку открытия шлагбаума '%s'",
        callback.from_user.id,
        barrier_key,
    )
    await callback.answer(f"Открываю {barrier['label']}…")
    await _open_barrier_and_report(callback.message, state, barrier_key)


@router.message(StateFilter(AdminStates.waiting_for_barrier_pin))
async def admin_barrier_pin(message: Message, state: FSMContext) -> None:
    if not is_private_message(message):
        return
    if not await is_admin(message.from_user.id):
        return

    data = await state.get_data()
    barrier_key = data.get("barrier_pin_key")
    pin = (message.text or "").strip()
    await state.clear()

    if not barrier_key:
        return
    logger.info(
        "Получен PIN от админа %s для шлагбаума '%s', повторяем логин",
        message.from_user.id,
        barrier_key,
    )
    await _open_barrier_and_report(message, state, barrier_key, pin=pin)


# ── Admin: FAQ ──────────────────────────────────────────────────

@router.callback_query(F.data == "admin:faq")
async def admin_faq_menu(callback: CallbackQuery) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(
        "❓ <b>Ответы на частые вопросы</b>\nВыберите раздел для редактирования:",
        reply_markup=faq_admin_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("faq:edit:"))
async def admin_faq_edit(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    key = callback.data.split(":")[2]
    label = "Шлагбаум" if key == "barrier" else "Бот"
    await state.update_data(faq_key=key)
    await state.set_state(AdminStates.waiting_for_faq_text)
    await callback.message.answer(f"Введите текст ответа для раздела <b>{label}</b>:")
    await callback.answer()


@router.message(StateFilter(AdminStates.waiting_for_faq_text))
async def admin_faq_save(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message) or not await is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text:
        await message.answer("Текст не может быть пустым.")
        return
    data = await state.get_data()
    key = data["faq_key"]
    await app.storage.set_faq(key, text)
    await state.clear()
    label = "Шлагбаум" if key == "barrier" else "Бот"
    await message.answer(
        f"✅ Ответ для раздела <b>{label}</b> сохранён.",
        reply_markup=await build_main_menu(message.from_user.id),
    )


# ── User: FAQ ───────────────────────────────────────────────────

@router.message(F.text == FAQ_BTN)
async def user_faq_menu(message: Message) -> None:
    if not is_private_message(message):
        return
    await message.answer("❓ <b>Ответы на частые вопросы</b>\nВыберите раздел:", reply_markup=faq_user_keyboard())


@router.callback_query(F.data.startswith("faq:get:"))
async def user_faq_get(callback: CallbackQuery) -> None:
    assert app is not None
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    key = callback.data.split(":")[2]
    text = await app.storage.get_faq(key)
    label = "Шлагбаум" if key == "barrier" else "Бот"
    if not text:
        await callback.answer(f"Ответ для раздела «{label}» ещё не добавлен администратором.", show_alert=True)
        return
    await callback.message.answer(f"<b>{label}</b>\n\n{text}")
    await callback.answer()


# ── AI Assistant ────────────────────────────────────────────────

AI_TOPICS = {"barrier": "🚗 Шлагбаум", "bot": "🤖 Бот", "general": "общий"}
AI_HISTORY_LIMIT = 5
AI_MENU_BUTTONS = {
    VERIFY_BTN,
    DELETE_VERIFY_BTN,
    ADMIN_BTN,
    DOCS_BTN,
    FAQ_BTN,
    AI_BTN,
}


def _ai_history_from_data(data: dict) -> list[tuple[str, str]]:
    raw = data.get("ai_history") or []
    return [(str(item[0]), str(item[1])) for item in raw if len(item) == 2]


async def _ai_start_topic(message: Message, state: FSMContext, topic: str, prompt: str) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    if not await app.storage.get_verification(message.from_user.id):
        await message.answer("Эта функция доступна только верифицированным пользователям.")
        return
    if not app.ai:
        await message.answer("AI-ассистент временно недоступен.")
        return
    await state.update_data(ai_topic=topic, ai_history=[])
    await state.set_state(AdminStates.waiting_for_ai_question)
    await message.answer(prompt)


async def _ai_reply(
    message: Message,
    state: FSMContext,
    *,
    question: str,
    topic: str,
    history: list[tuple[str, str]],
) -> None:
    assert app is not None
    thinking = await message.answer("⏳ Ищу ответ...")
    answer = await app.ai.ask(topic, question, history=history)
    log_user_ai_exchange(
        user_id=message.from_user.id,
        username=message.from_user.username,
        full_name=message.from_user.full_name or "",
        topic=topic,
        question=question,
        answer=answer,
    )
    await thinking.delete()

    updated_history = history + [[question, answer]]
    if len(updated_history) > AI_HISTORY_LIMIT:
        updated_history = updated_history[-AI_HISTORY_LIMIT:]
    await state.update_data(ai_history=updated_history)
    await state.set_state(AdminStates.waiting_for_ai_choice)
    await message.answer(answer, reply_markup=ai_conversation_keyboard())


@router.message(F.text == AI_BTN)
async def ai_start(message: Message, state: FSMContext) -> None:
    await _ai_start_topic(message, state, "general", "Задайте ваш вопрос:")


@router.message(StateFilter(AdminStates.waiting_for_ai_question))
async def ai_answer(message: Message, state: FSMContext) -> None:
    assert app is not None
    if not is_private_message(message):
        return
    question = (message.text or "").strip()
    if not question:
        await message.answer("Пожалуйста, введите вопрос текстом.")
        return
    if question in AI_MENU_BUTTONS:
        await state.clear()
        await message.answer(
            "Разговор с помощником завершён.",
            reply_markup=await build_main_menu(message.from_user.id),
        )
        return

    data = await state.get_data()
    topic = data.get("ai_topic", "general")
    history = _ai_history_from_data(data)
    await _ai_reply(message, state, question=question, topic=topic, history=history)


@router.message(StateFilter(AdminStates.waiting_for_ai_choice))
async def ai_choice_prompt(message: Message) -> None:
    if not is_private_message(message):
        return
    await message.answer(
        "Выберите «Продолжить разговор» или «Завершить разговор».",
        reply_markup=ai_conversation_keyboard(),
    )


@router.callback_query(F.data == "ai:continue")
async def ai_continue(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return
    current_state = await state.get_state()
    if current_state != AdminStates.waiting_for_ai_choice.state:
        await callback.answer("Сессия помощника уже завершена.", show_alert=True)
        return

    await callback.answer()
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            logger.debug("Could not remove AI choice keyboard", exc_info=True)

    await state.set_state(AdminStates.waiting_for_ai_question)
    if callback.message:
        await callback.message.answer(
            "Задайте дополнительный вопрос:",
            reply_markup=ReplyKeyboardRemove(),
        )


@router.callback_query(F.data == "ai:end")
async def ai_end(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return

    await callback.answer()
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            logger.debug("Could not remove AI choice keyboard", exc_info=True)

    await state.clear()
    if callback.message:
        await callback.message.answer(
            "Разговор с помощником завершён.",
            reply_markup=await build_main_menu(callback.from_user.id),
        )


@router.message()
async def fallback(message: Message) -> None:
    if not is_private_message(message):
        return
    await message.answer("Используйте кнопки меню или команду /start.", reply_markup=await build_main_menu(message.from_user.id))



@router.callback_query(F.data == "admin:export_unverified")
async def admin_export_unverified(callback: CallbackQuery) -> None:
    assert app is not None

    if not is_private_callback(callback):
        await reject_group_callback(callback)
        return

    if not await is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.answer()
    await callback.message.answer("Собираю список неверифицированных участников, подождите...")

    try:
        await sync_group_members()
    except Exception:
        logger.exception("Failed to sync group members for unverified export")
        await callback.message.answer("Не удалось синхронизировать участников группы.")
        return

    users = await get_unverified_action_users()

    if not users:
        await callback.message.answer("Неверифицированных участников нет.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Неверифицированные"

    ws.append(["Идентификатор", "Имя пользователя", "Полное имя", "Первый раз в группе"])

    for user in users:
        profile = await app.storage.get_user_profile(user["user_id"])
        first_seen = profile.get("first_seen_at", "") if profile else ""
        username, full_name = await get_live_display_name(user["user_id"])
        ws.append([
            user["user_id"],
            username or "",
            full_name,
            first_seen,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    document = BufferedInputFile(
        buffer.read(),
        filename="unverified_users.xlsx"
    )
    await callback.message.answer(f"Неверифицированных участников: {len(users)}")
    await callback.message.answer_document(document)

async def on_startup() -> None:
    assert app is not None
    await app.storage.init()
    await app.storage.seed_admins(app.settings.admin_ids)
    me = await app.bot.get_me()
    app.bot_user_id = me.id
    logger.info("Bot user id cached for membership stats: %s", app.bot_user_id)
    # Init DeepSeek RAG
    knowledge_dir = os.path.join(os.path.dirname(__file__), "..", "knowledge")
    rag_dir = os.path.join(os.path.dirname(__file__), "..", "rag_store")
    if app.settings.deepseek_api_key:
        app.ai = DeepSeekClient(app.settings.deepseek_api_key, knowledge_dir, rag_dir)
        app.ai.reload_knowledge()
        logger.info("DeepSeek RAG client initialized")
    else:
        logger.warning("DEEPSEEK_API_KEY not set, AI assistant disabled")

    # Sync documents: remove DB entries whose files no longer exist
    docs_dir = os.path.join(os.path.dirname(__file__), "..", "documents")
    docs = await app.storage.list_documents()
    for doc in docs:
        file_path = os.path.join(docs_dir, doc["filename"])
        if not os.path.exists(file_path):
            logger.warning("Document file missing, removing from DB: %s (%s)", doc["name"], doc["filename"])
            await app.storage.delete_document(doc["doc_id"])
    await app.telethon.start()
    await sync_group_members()
    logger.info("Application started")


async def on_shutdown() -> None:
    assert app is not None
    await app.telethon.disconnect()
    await app.bot.session.close()
    logger.info("Application stopped")


async def run() -> None:
    global app
    settings = load_settings()
    log_level = getattr(logging, settings.log_level.upper(), logging.INFO)
    log_fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s")

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_fmt)

    # File handler — new file every day, keep 30 days
    log_dir = os.getenv("LOG_DIR", "logs").strip() or "logs"
    os.makedirs(log_dir, exist_ok=True)
    file_handler = TimedRotatingFileHandler(
        filename=os.path.join(log_dir, "house_bot.log"),
        when="midnight",
        interval=1,
        backupCount=30,
        encoding="utf-8",
        utc=True,
    )
    file_handler.suffix = "%Y-%m-%d"
    file_handler.setFormatter(log_fmt)

    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)
    root_logger.addHandler(console_handler)
    root_logger.addHandler(file_handler)
    setup_prompt_logger(log_dir)
    app = AppContext(settings)
    await on_startup()
    try:
        await app.dp.start_polling(app.bot)
    finally:
        await on_shutdown()
